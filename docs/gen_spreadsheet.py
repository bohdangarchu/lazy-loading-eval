import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

MODELS = [
    ("openai-community/gpt2",        "548 MB",  "python:3.12-slim",      "43.2 MB"),
    ("facebook/opt-350m",            "663 MB",  "tensorflow/tensorflow", "588 MB"),
    ("Qwen/Qwen2-1.5B",              "3.09 GB", "ollama/ollama",         "3.76 GB"),
    ("openlm-research/open_llama_3b","6.85 GB", "ollama/ollama",         "3.76 GB"),
]

ALL_MODES     = ["2dfs", "2dfs-stargz", "2dfs-stargz-zstd", "stargz", "base"]
REFRESH_MODES = ["2dfs-stargz", "2dfs-stargz-zstd"]
ENV           = "Cloudlab, 2 c6526-25g nodes"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF")
ALT_FILL    = PatternFill("solid", fgColor="D9E1F2")


ENV_LABEL_FONT = Font(bold=True)


def write_env_row(ws):
    ws.append(["Environment:", ENV])
    ws["A1"].font = ENV_LABEL_FONT
    ws.append([])  # blank spacer


def style_header(ws, row):
    for cell in ws[row]:
        if cell.value is not None:
            cell.fill      = HEADER_FILL
            cell.font      = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")


def autofit(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 45)


def add_validation(ws, col_letter, first_row, last_row):
    from openpyxl.worksheet.datavalidation import DataValidation
    dv = DataValidation(type="list", formula1='"TODO,In Progress,Done"', allow_blank=True)
    dv.sqref = f"{col_letter}{first_row}:{col_letter}{last_row}"
    ws.add_data_validation(dv)


def append_rows(ws, rows, status_col, header_row):
    for i, row in enumerate(rows):
        ws.append(row)
        if i % 2 == 1:
            for cell in ws[ws.max_row]:
                cell.fill = ALT_FILL
    add_validation(ws, status_col, header_row + 1, header_row + len(rows))


def write_sheet_build(wb):
    ws = wb.create_sheet("Build Performance")
    write_env_row(ws)
    ws.append(["Model", "Model Size", "Base Image", "Base Image Size", "Mode", "Splits", "n_runs", "Status"])
    style_header(ws, 3)

    rows = [
        [model, msize, base, bsize, mode, "1-10", 3, "Done"]
        for model, msize, base, bsize in MODELS
        for mode in ALL_MODES
    ]
    append_rows(ws, rows, "H", 3)
    autofit(ws)


def write_sheet_rebuild(wb):
    ws = wb.create_sheet("Rebuild Performance")
    write_env_row(ws)
    ws.append(["Model", "Model Size", "Base Image", "Base Image Size", "Mode", "Splits (fixed)", "r (chunks mutated)", "Directions", "n_runs", "Status"])
    style_header(ws, 3)

    rows = [
        [model, msize, base, bsize, mode, 10, "[2, 4, 6, 8, 10]", "top_to_bottom, bottom_to_top", 3, "Done"]
        for model, msize, base, bsize in MODELS
        for mode in ALL_MODES
    ]
    append_rows(ws, rows, "J", 3)
    autofit(ws)


def write_sheet_pull(wb):
    ws = wb.create_sheet("Pull Performance")
    write_env_row(ws)
    ws.append(["Model", "Model Size", "Base Image", "Base Image Size", "Mode", "Splits", "n_runs", "Status"])
    style_header(ws, 3)

    rows = [
        [model, msize, base, bsize, mode, "[2, 4, 6, 8, 10]", 3, "Done"]
        for model, msize, base, bsize in MODELS
        for mode in ALL_MODES
    ]
    append_rows(ws, rows, "H", 3)
    autofit(ws)


def write_sheet_refresh(wb):
    ws = wb.create_sheet("Layer Refresh")
    write_env_row(ws)
    ws.append(["Model", "Model Size", "Base Image", "Base Image Size", "Mode", "k (layers refreshed)", "n_runs", "Status"])
    style_header(ws, 3)

    rows = [
        [model, msize, base, bsize, mode, "1-10", 3, "Done"]
        for model, msize, base, bsize in MODELS
        for mode in REFRESH_MODES
    ]
    append_rows(ws, rows, "H", 3)
    autofit(ws)


wb = openpyxl.Workbook()
wb.remove(wb.active)
write_sheet_build(wb)
write_sheet_rebuild(wb)
write_sheet_pull(wb)
write_sheet_refresh(wb)

out = "/home/garchu/workspace/lazy-loading-eval/experiments.xlsx"
wb.save(out)
print(f"Saved: {out}")
